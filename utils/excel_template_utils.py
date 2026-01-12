"""
أدوات قوالب Excel للاستيراد والتصدير
Excel Template Utilities for Import/Export
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from typing import List, Dict, Any
from translations import get_language


class ExcelTemplateGenerator:
    """مولد قوالب Excel"""

    def __init__(self, language='ar'):
        self.language = language
        self.is_rtl = (language == 'ar')

    def _style_header(self, ws, row, columns):
        """تنسيق صف العنوان"""
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_num, col_name in enumerate(columns, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = col_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            ws.column_dimensions[get_column_letter(col_num)].width = 20

    def generate_clos_template(self, output_path):
        """
        إنشاء قالب Excel لمخرجات التعلم (CLOs)

        Args:
            output_path: مسار حفظ الملف
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CLOs" if self.language == 'en' else "مخرجات التعلم"

        # العناوين
        if self.language == 'ar':
            columns = ['رقم المخرج', 'التصنيف', 'وصف المخرج', 'PLOs المرتبطة']
        else:
            columns = ['CLO Code', 'Category', 'CLO Description', 'Aligned PLOs']

        self._style_header(ws, 1, columns)

        # إضافة أمثلة
        examples = [
            ['1', 'Knowledge', 'يحدد الطالب المفاهيم الأساسية للإحصاء', 'K1, K2'],
            ['2', 'Skills', 'يطبق الطالب الأساليب الإحصائية', 'S1, S3'],
            ['3', 'Values', 'يقدر الطالب أهمية الدقة في التحليل الإحصائي', 'V1'],
        ]

        for row_num, example in enumerate(examples, 2):
            for col_num, value in enumerate(example, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # ملاحظات
        notes_row = len(examples) + 4
        ws.cell(row=notes_row, column=1).value = "ملاحظات:" if self.language == 'ar' else "Notes:"
        ws.cell(row=notes_row, column=1).font = Font(bold=True)
        ws.cell(row=notes_row+1, column=1).value = "- احذف الأمثلة وأدخل بياناتك" if self.language == 'ar' else "- Delete examples and enter your data"
        ws.cell(row=notes_row+2, column=1).value = "- رقم المخرج: أي رقم أو كود (مثل: 1, 2, CLO1)" if self.language == 'ar' else "- CLO Code: any number or code (e.g., 1, 2, CLO1)"
        ws.cell(row=notes_row+3, column=1).value = "- التصنيف: Knowledge أو Skills أو Values فقط" if self.language == 'ar' else "- Category: Knowledge, Skills, or Values only"
        ws.cell(row=notes_row+4, column=1).value = "- PLOs المرتبطة: اختياري (مثل: K1, K2)" if self.language == 'ar' else "- Aligned PLOs: optional (e.g., K1, K2)"

        wb.save(output_path)
        return True

    def generate_topics_template(self, output_path):
        """
        إنشاء قالب Excel للمواضيع

        Args:
            output_path: مسار حفظ الملف
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Topics" if self.language == 'en' else "المواضيع"

        # العناوين
        if self.language == 'ar':
            columns = ['رقم الموضوع', 'عنوان الموضوع', 'عدد الساعات']
        else:
            columns = ['Topic Number', 'Topic Title', 'Contact Hours']

        self._style_header(ws, 1, columns)

        # إضافة أمثلة
        examples = [
            ['1', 'الإحصاء الوصفي', '3'],
            ['2', 'الاحتمالات', '4'],
            ['3', 'التوزيعات الاحتمالية', '5'],
        ]

        for row_num, example in enumerate(examples, 2):
            for col_num, value in enumerate(example, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # ملاحظات
        notes_row = len(examples) + 4
        ws.cell(row=notes_row, column=1).value = "ملاحظات:" if self.language == 'ar' else "Notes:"
        ws.cell(row=notes_row, column=1).font = Font(bold=True)
        ws.cell(row=notes_row+1, column=1).value = "- احذف الأمثلة وأدخل بياناتك" if self.language == 'ar' else "- Delete examples and enter your data"
        ws.cell(row=notes_row+2, column=1).value = "- رقم الموضوع: يجب أن يكون رقماً صحيحاً (1, 2, 3...)" if self.language == 'ar' else "- Topic Number: must be an integer (1, 2, 3...)"
        ws.cell(row=notes_row+3, column=1).value = "- عنوان الموضوع: يمكن كتابته بأي لغة" if self.language == 'ar' else "- Topic Title: can be written in any language"
        ws.cell(row=notes_row+4, column=1).value = "- عدد الساعات: يجب أن يكون رقماً (مثل: 3, 4.5)" if self.language == 'ar' else "- Contact Hours: must be a number (e.g., 3, 4.5)"

        wb.save(output_path)
        return True

    def generate_activities_template(self, output_path):
        """
        إنشاء قالب Excel للأنشطة

        Args:
            output_path: مسار حفظ الملف
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Activities" if self.language == 'en' else "الأنشطة"

        # العناوين
        if self.language == 'ar':
            columns = ['رقم النشاط', 'اسم النشاط', 'الدرجة', 'النسبة المئوية', 'التوقيت', 'مخرجات التعلم المرتبطة']
        else:
            columns = ['Activity Number', 'Activity Name', 'Mark', 'Percentage', 'Timing', 'Link to CLOs']

        self._style_header(ws, 1, columns)

        # ضبط عرض الأعمدة
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 25

        # إضافة أمثلة
        examples = [
            ['1', 'اختبار منتصف الفصل', '30', '30', 'الأسبوع 8', '1, 2'],
            ['2', 'الاختبار النهائي', '40', '40', 'نهاية الفصل', '1, 2, 3'],
            ['3', 'مشروع', '15', '15', 'الأسبوع 12', '2, 3'],
            ['4', 'واجبات', '10', '10', 'أسبوعي', '1'],
            ['5', 'مشاركة', '5', '5', 'طوال الفصل', '1, 2, 3'],
        ]

        for row_num, example in enumerate(examples, 2):
            for col_num, value in enumerate(example, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # ملاحظات
        notes_row = len(examples) + 4
        ws.cell(row=notes_row, column=1).value = "ملاحظات:" if self.language == 'ar' else "Notes:"
        ws.cell(row=notes_row, column=1).font = Font(bold=True)
        ws.cell(row=notes_row+1, column=1).value = "- احذف الأمثلة وأدخل بياناتك" if self.language == 'ar' else "- Delete examples and enter your data"
        ws.cell(row=notes_row+2, column=1).value = "- رقم النشاط: اختياري، يمكن تركه فارغاً" if self.language == 'ar' else "- Activity Number: optional, can be left empty"
        ws.cell(row=notes_row+3, column=1).value = "- اسم النشاط: مطلوب (مثل: اختبار منتصف الفصل)" if self.language == 'ar' else "- Activity Name: required (e.g., Midterm Exam)"
        ws.cell(row=notes_row+4, column=1).value = "- الدرجة: رقم (مثل: 30, 40)" if self.language == 'ar' else "- Mark: number (e.g., 30, 40)"
        ws.cell(row=notes_row+5, column=1).value = "- النسبة المئوية: رقم بدون علامة % (مثل: 30, 40)" if self.language == 'ar' else "- Percentage: number without % sign (e.g., 30, 40)"
        ws.cell(row=notes_row+6, column=1).value = "- التوقيت: اختياري (مثل: الأسبوع 8، نهاية الفصل)" if self.language == 'ar' else "- Timing: optional (e.g., Week 8, End of semester)"
        ws.cell(row=notes_row+7, column=1).value = "- مخرجات التعلم: أرقام المخرجات مفصولة بفاصلة (مثل: 1, 2, 3)" if self.language == 'ar' else "- Link to CLOs: CLO numbers separated by comma (e.g., 1, 2, 3)"

        wb.save(output_path)
        return True


class ExcelTemplateReader:
    """قارئ قوالب Excel"""

    @staticmethod
    def read_clos_from_excel(file_path) -> List[Dict[str, Any]]:
        """
        قراءة مخرجات التعلم من ملف Excel

        Args:
            file_path: مسار ملف Excel

        Returns:
            قائمة بمخرجات التعلم
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        clos = []
        # البدء من الصف 2 (تخطي العنوان)
        for row in ws.iter_rows(min_row=2, values_only=True):
            # تخطي الصفوف الفارغة أو التي تحتوي على ملاحظات
            if not row[0] or str(row[0]).strip() == '' or str(row[0]).startswith('-'):
                continue

            try:
                # الأعمدة الجديدة: رقم المخرج | التصنيف | وصف المخرج | PLOs المرتبطة
                clo = {
                    'code': str(row[0]).strip(),  # رقم/كود المخرج
                    'category': str(row[1]).strip() if row[1] else 'Knowledge',  # التصنيف
                    'description': str(row[2]).strip() if row[2] else '',  # الوصف
                    'aligned_plos': str(row[3]).strip() if row[3] else ''  # PLOs المرتبطة
                }
                clos.append(clo)
            except (ValueError, TypeError, IndexError) as e:
                # تخطي الصفوف التي بها أخطاء
                continue

        return clos

    @staticmethod
    def read_topics_from_excel(file_path) -> List[Dict[str, Any]]:
        """
        قراءة المواضيع من ملف Excel

        Args:
            file_path: مسار ملف Excel

        Returns:
            قائمة بالمواضيع
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        topics = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0] or str(row[0]).strip() == '' or str(row[0]).startswith('-'):
                continue

            try:
                # الأعمدة الجديدة: رقم الموضوع | عنوان الموضوع | عدد الساعات
                topic = {
                    'number': int(row[0]),
                    'title': str(row[1]).strip() if row[1] else '',
                    'hours': float(row[2]) if row[2] else 0
                }
                topics.append(topic)
            except (ValueError, TypeError, IndexError):
                continue

        return topics

    @staticmethod
    def read_activities_from_excel(file_path) -> List[Dict[str, Any]]:
        """
        قراءة الأنشطة من ملف Excel

        Args:
            file_path: مسار ملف Excel

        Returns:
            قائمة بالأنشطة
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        activities = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # تخطي الصفوف الفارغة أو التي تحتوي على ملاحظات
            # نتحقق من العمود الثاني (اسم النشاط) لأن رقم النشاط اختياري
            if not row[1] or str(row[1]).strip() == '' or str(row[1]).startswith('-'):
                continue

            try:
                # الأعمدة: رقم النشاط | اسم النشاط | الدرجة | النسبة المئوية | التوقيت | مخرجات التعلم
                activity = {
                    'number': int(row[0]) if row[0] and str(row[0]).strip() != '' else None,
                    'name': str(row[1]).strip() if row[1] else '',
                    'mark': float(row[2]) if row[2] else 0,
                    'percentage': float(row[3]) if row[3] else 0,
                    'timing': str(row[4]).strip() if row[4] else '',
                    'linked_clos': str(row[5]).strip() if row[5] else ''
                }
                activities.append(activity)
            except (ValueError, TypeError, IndexError):
                continue

        return activities
