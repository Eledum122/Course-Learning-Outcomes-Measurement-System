"""
تقرير متابعة تقدم الشعب
Section Progress Tracking Report
"""

import streamlit as st
import pandas as pd
import json
from pathlib import Path
from datetime import datetime
from io import BytesIO
from models.database import Database, UserRole
from utils.permissions import get_permissions_helper

# Try to import openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def load_sections_data():
    """Load sections data"""
    sections_file = Path(__file__).parent.parent / 'data' / 'sections.json'
    if sections_file.exists():
        try:
            with open(sections_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {"sections": []}
    return {"sections": []}


def load_programs_data():
    """Load programs data"""
    programs_file = Path(__file__).parent.parent / 'data' / 'programs.json'
    if programs_file.exists():
        try:
            with open(programs_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {"programs": []}
    return {"programs": []}


def load_courses_data():
    """Load courses data"""
    courses_file = Path(__file__).parent.parent / 'data' / 'courses.json'
    if courses_file.exists():
        try:
            with open(courses_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {"courses": []}
    return {"courses": []}


def load_section_students_data():
    """Load section students data"""
    students_file = Path(__file__).parent.parent / 'data' / 'section_students.json'
    if students_file.exists():
        try:
            with open(students_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {"section_students": []}
    return {"section_students": []}


def load_student_grades_data():
    """Load student grades data"""
    grades_file = Path(__file__).parent.parent / 'data' / 'student_grades.json'
    if grades_file.exists():
        try:
            with open(grades_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {"student_grades": []}
    return {"student_grades": []}


def load_faculty_data():
    """Load faculty data"""
    faculty_file = Path(__file__).parent.parent / 'data' / 'faculty.json'
    if faculty_file.exists():
        try:
            with open(faculty_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {"faculty_members": []}
    return {"faculty_members": []}


def get_instructor_name(instructor_id, faculty_members):
    """Get instructor name from ID"""
    if not instructor_id:
        return "غير محدد / Not Assigned"
    for member in faculty_members:
        if member.get('employee_id') == instructor_id:
            return member.get('name', '')
    return instructor_id


def get_section_students_count(section_id, all_students):
    """Get number of students in a section"""
    return len([s for s in all_students if s.get('section_id') == section_id])


def get_section_regular_students(section_id, all_students):
    """Get regular students (not dropped) in a section"""
    return [s for s in all_students if s.get('section_id') == section_id and s.get('status') == 'Regular']


def get_course_activities(course_data):
    """Get assessment activities for a course with their CLO requirements"""
    activities = {}
    clos_activities_distribution = course_data.get('clos_activities_distribution', {})

    for activity_name, clo_marks in clos_activities_distribution.items():
        # Get CLOs that have non-zero marks (i.e., are assessed in this activity)
        required_clos = [clo_code for clo_code, mark in clo_marks.items() if mark and mark > 0]
        if required_clos:  # Only include activities with at least one CLO
            activities[activity_name] = required_clos

    return activities


def get_grades_progress_for_section(section_id, activities_with_clos, all_students, all_grades):
    """
    Calculate grade entry progress for each activity in a section

    An activity is considered complete for a student if ALL required CLO grades
    have been entered for that activity.

    Args:
        section_id: The section ID
        activities_with_clos: Dict of {activity_name: [list of required CLO codes]}
        all_students: List of all students
        all_grades: List of all grades
    """
    regular_students = get_section_regular_students(section_id, all_students)

    if not regular_students:
        return {}

    progress = {}
    section_grades = [g for g in all_grades if g.get('section_id') == section_id]

    # Build a lookup: {student_record_id: {(activity_name, clo_code): mark}}
    grades_lookup = {}
    for grade in section_grades:
        student_id = grade.get('student_record_id')
        activity_name = grade.get('activity_name')
        clo_code = grade.get('clo_code')
        mark = grade.get('mark')

        if student_id not in grades_lookup:
            grades_lookup[student_id] = {}
        grades_lookup[student_id][(activity_name, clo_code)] = mark

    for activity_name, required_clos in activities_with_clos.items():
        # Count students who have ALL required CLO grades for this activity
        students_complete = 0

        for student in regular_students:
            student_id = student.get('student_record_id')
            student_grades = grades_lookup.get(student_id, {})

            # Check if all required CLOs have grades
            all_clos_entered = True
            for clo_code in required_clos:
                grade_key = (activity_name, clo_code)
                if grade_key not in student_grades or student_grades[grade_key] is None:
                    all_clos_entered = False
                    break

            if all_clos_entered:
                students_complete += 1

        total = len(regular_students)
        percentage = (students_complete / total * 100) if total > 0 else 0

        progress[activity_name] = {
            'entered': students_complete,
            'total': total,
            'percentage': percentage,
            'required_clos': len(required_clos)
        }

    return progress


def get_progress_color(percentage):
    """Get color based on progress percentage"""
    if percentage >= 100:
        return "#28a745"  # Green
    elif percentage >= 75:
        return "#17a2b8"  # Blue
    elif percentage >= 50:
        return "#ffc107"  # Yellow
    elif percentage > 0:
        return "#fd7e14"  # Orange
    else:
        return "#dc3545"  # Red


def get_progress_icon(percentage):
    """Get icon based on progress percentage"""
    if percentage >= 100:
        return "✅"
    elif percentage >= 75:
        return "🔵"
    elif percentage >= 50:
        return "🟡"
    elif percentage > 0:
        return "🟠"
    else:
        return "🔴"


def export_to_excel(sections_data, course_code, activities_with_clos):
    """Export progress report to Excel"""
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Section Progress"

    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    activities = list(activities_with_clos.keys())

    # Build headers
    headers = ['Section', 'Instructor', 'Students', 'Students %']
    for activity in activities:
        clos_count = len(activities_with_clos.get(activity, []))
        headers.append(f'{activity} ({clos_count} CLOs) Count')
        headers.append(f'{activity} (%)')
    headers.append('Overall %')

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Write data
    for row_idx, section in enumerate(sections_data, 2):
        col = 1

        # Section number
        ws.cell(row=row_idx, column=col, value=section['section_number']).border = thin_border
        col += 1

        # Instructor
        ws.cell(row=row_idx, column=col, value=section['instructor']).border = thin_border
        col += 1

        # Students count
        ws.cell(row=row_idx, column=col, value=section['students_count']).border = thin_border
        col += 1

        # Students percentage (assuming expected is same as enrolled for now)
        cell = ws.cell(row=row_idx, column=col, value=f"{section['students_percentage']:.0f}%")
        cell.border = thin_border
        col += 1

        # Activity progress
        for activity in activities:
            progress = section['activities_progress'].get(activity, {})

            # Count
            count_str = f"{progress.get('entered', 0)}/{progress.get('total', 0)}"
            ws.cell(row=row_idx, column=col, value=count_str).border = thin_border
            col += 1

            # Percentage
            pct = progress.get('percentage', 0)
            cell = ws.cell(row=row_idx, column=col, value=f"{pct:.0f}%")
            cell.border = thin_border

            # Color coding
            if pct >= 100:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif pct >= 50:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif pct > 0:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            col += 1

        # Overall percentage
        cell = ws.cell(row=row_idx, column=col, value=f"{section['overall_percentage']:.0f}%")
        cell.border = thin_border
        if section['overall_percentage'] >= 100:
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif section['overall_percentage'] >= 50:
            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output)
    output.seek(0)
    return output.getvalue()


def show_section_progress_report(db: Database, user, lang: str):
    """Display Section Progress Tracking Report"""

    st.title("📋 Section Progress Report / تقرير متابعة تقدم الشعب")
    st.caption("Track student enrollment and grade entry progress for each section")

    # Initialize permissions helper
    perm = get_permissions_helper(db, user)

    # Load data with permissions filtering
    sections_data = load_sections_data()
    all_sections = sections_data.get('sections', [])
    sections = perm.filter_sections(all_sections)

    programs_data = load_programs_data()
    all_programs = programs_data.get('programs', [])
    programs = perm.filter_programs(all_programs)

    courses_data = load_courses_data()
    all_courses = courses_data.get('courses', [])
    courses = perm.filter_courses(all_courses)

    students_data = load_section_students_data()
    all_students = students_data.get('section_students', [])

    grades_data = load_student_grades_data()
    all_grades = grades_data.get('student_grades', [])

    faculty_data = load_faculty_data()
    faculty_members = faculty_data.get('faculty_members', [])

    if not sections:
        st.warning("لا توجد شعب متاحة / No sections available")
        return

    # ===============================
    # Section 1: Hierarchical Selection
    # ===============================
    st.markdown("### 1. Select Course and Semester / اختر المقرر والفصل")

    # Row 1: Program and Course
    col1, col2 = st.columns(2)

    with col1:
        program_options = ["-- Select Program / اختر البرنامج --"]
        program_map = {}
        for p in programs:
            if p.get('is_active', True):
                display = f"{p.get('program_code', '')} - {p.get('program_name_en', '')}"
                program_options.append(display)
                program_map[display] = p.get('program_id')

        selected_program_display = st.selectbox(
            "Program / البرنامج",
            program_options,
            key="progress_program_filter"
        )

        selected_program_id = None
        if selected_program_display != "-- Select Program / اختر البرنامج --":
            selected_program_id = program_map.get(selected_program_display)

    with col2:
        if selected_program_id:
            program_courses = [c for c in courses if c.get('program_id') == selected_program_id]
            course_options = ["-- Select Course / اختر المقرر --"]
            course_map = {}
            for c in program_courses:
                display = f"{c.get('course_code', '')} - {c.get('course_title_en', '')}"
                course_options.append(display)
                course_map[display] = c.get('course_id')

            selected_course_display = st.selectbox(
                "Course / المقرر",
                course_options,
                key="progress_course_filter"
            )

            selected_course_id = None
            if selected_course_display != "-- Select Course / اختر المقرر --":
                selected_course_id = course_map.get(selected_course_display)
        else:
            st.selectbox(
                "Course / المقرر",
                ["-- Select Program First / اختر البرنامج أولاً --"],
                disabled=True,
                key="progress_course_disabled"
            )
            selected_course_id = None

    # Row 2: Academic Year and Semester
    col3, col4 = st.columns(2)

    with col3:
        if selected_course_id:
            course_sections = [s for s in sections if s.get('course_id') == selected_course_id]
            available_years = sorted(list(set(s.get('academic_year', '') for s in course_sections if s.get('academic_year'))), reverse=True)
            year_options = ["-- All Years / كل السنوات --"] + available_years

            selected_year = st.selectbox(
                "Academic Year / السنة الأكاديمية",
                year_options,
                key="progress_year_filter"
            )

            if selected_year == "-- All Years / كل السنوات --":
                selected_year = None
        else:
            st.selectbox(
                "Academic Year / السنة الأكاديمية",
                ["-- Select Course First / اختر المقرر أولاً --"],
                disabled=True,
                key="progress_year_disabled"
            )
            selected_year = None

    with col4:
        if selected_course_id:
            if selected_year:
                year_sections = [s for s in course_sections if s.get('academic_year') == selected_year]
            else:
                year_sections = course_sections

            available_semesters = sorted(list(set(s.get('semester', '') for s in year_sections if s.get('semester'))))
            semester_options = ["-- All Semesters / كل الفصول --"] + available_semesters

            selected_semester = st.selectbox(
                "Semester / الفصل الدراسي",
                semester_options,
                key="progress_semester_filter"
            )

            if selected_semester == "-- All Semesters / كل الفصول --":
                selected_semester = None
        else:
            st.selectbox(
                "Semester / الفصل الدراسي",
                ["-- Select Course First / اختر المقرر أولاً --"],
                disabled=True,
                key="progress_semester_disabled"
            )
            selected_semester = None

    # Check if course is selected
    if not selected_course_id:
        st.info("💡 Please select a program and course to view the progress report")
        return

    # Get course data and activities
    course_data = next((c for c in courses if c.get('course_id') == selected_course_id), None)
    if not course_data:
        st.error("Course data not found")
        return

    activities_with_clos = get_course_activities(course_data)
    activities = list(activities_with_clos.keys())  # Activity names list for display

    # Filter sections
    filtered_sections = [s for s in sections if s.get('course_id') == selected_course_id]
    if selected_year:
        filtered_sections = [s for s in filtered_sections if s.get('academic_year') == selected_year]
    if selected_semester:
        filtered_sections = [s for s in filtered_sections if s.get('semester') == selected_semester]

    if not filtered_sections:
        st.warning("لا توجد شعب للمعايير المحددة / No sections found for the selected criteria")
        return

    st.markdown("---")

    # ===============================
    # Section 2: Overview Statistics
    # ===============================
    st.markdown("### 2. Overview / نظرة عامة")

    # Calculate overall statistics
    total_sections = len(filtered_sections)
    total_students = 0
    sections_with_students = 0
    sections_with_complete_grades = 0

    sections_report_data = []

    for section in filtered_sections:
        section_id = section.get('section_id')
        section_number = section.get('section_number', '')
        instructor_id = section.get('instructor_id', '')
        instructor_name = get_instructor_name(instructor_id, faculty_members)

        # Students count
        students_count = get_section_students_count(section_id, all_students)
        regular_students = get_section_regular_students(section_id, all_students)
        regular_count = len(regular_students)

        total_students += students_count
        if students_count > 0:
            sections_with_students += 1

        # Students percentage (100% if has students, 0% if not)
        students_percentage = 100 if students_count > 0 else 0

        # Grades progress
        activities_progress = get_grades_progress_for_section(section_id, activities_with_clos, all_students, all_grades)

        # Calculate overall grade progress
        if activities and regular_count > 0:
            total_percentage = sum(p.get('percentage', 0) for p in activities_progress.values())
            overall_percentage = total_percentage / len(activities) if activities else 0
        else:
            overall_percentage = 0

        if overall_percentage >= 100:
            sections_with_complete_grades += 1

        sections_report_data.append({
            'section_id': section_id,
            'section_number': section_number,
            'instructor': instructor_name,
            'instructor_id': instructor_id,
            'students_count': students_count,
            'regular_count': regular_count,
            'students_percentage': students_percentage,
            'activities_progress': activities_progress,
            'overall_percentage': overall_percentage,
            'semester': section.get('semester', ''),
            'academic_year': section.get('academic_year', ''),
            'gender': section.get('gender', '')
        })

    # Display overview cards
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    padding: 20px; border-radius: 10px; color: white; text-align: center;">
            <h2 style="color: white; margin: 0;">📖 {total_sections}</h2>
            <p style="color: white; margin: 5px 0;">عدد الشعب / Total Sections</p>
        </div>
        """, unsafe_allow_html=True)

    with col2:
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
                    padding: 20px; border-radius: 10px; color: white; text-align: center;">
            <h2 style="color: white; margin: 0;">👨‍🎓 {total_students}</h2>
            <p style="color: white; margin: 5px 0;">إجمالي الطلاب / Total Students</p>
        </div>
        """, unsafe_allow_html=True)

    with col3:
        students_pct = (sections_with_students / total_sections * 100) if total_sections > 0 else 0
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%);
                    padding: 20px; border-radius: 10px; color: white; text-align: center;">
            <h2 style="color: white; margin: 0;">👥 {sections_with_students}/{total_sections}</h2>
            <p style="color: white; margin: 5px 0;">شعب بها طلاب / With Students</p>
        </div>
        """, unsafe_allow_html=True)

    with col4:
        grades_pct = (sections_with_complete_grades / total_sections * 100) if total_sections > 0 else 0
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #43e97b 0%, #38f9d7 100%);
                    padding: 20px; border-radius: 10px; color: white; text-align: center;">
            <h2 style="color: white; margin: 0;">✅ {sections_with_complete_grades}/{total_sections}</h2>
            <p style="color: white; margin: 5px 0;">مكتملة الدرجات / Complete Grades</p>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ===============================
    # Section 3: Detailed Progress Table
    # ===============================
    st.markdown("### 3. Detailed Progress / التقدم التفصيلي")

    if not activities:
        st.warning("لم يتم تحديد توزيع المخرجات على الأنشطة. يرجى إكمال المرحلة 2.4 أولاً / CLOs distribution not defined. Please complete Stage 2.4 first.")

    # Create section selector
    section_options = {
        f"📖 {s['section_number']} - {s['instructor']} ({get_progress_icon(s['overall_percentage'])} {s['overall_percentage']:.0f}%)": idx
        for idx, s in enumerate(sections_report_data)
    }

    if section_options:
        selected_section_display = st.selectbox(
            "اختر الشعبة لعرض التفاصيل / Select Section for Details",
            options=list(section_options.keys()),
            key="section_detail_selector"
        )

        selected_idx = section_options[selected_section_display]
        section = sections_report_data[selected_idx]

        # Section details card
        st.markdown(f"""
        <div style="background: #f8f9fa; padding: 20px; border-radius: 10px; margin: 10px 0; border: 1px solid #e0e0e0;">
            <div style="display: flex; flex-wrap: wrap; gap: 20px; justify-content: space-between;">
                <div><strong>الشعبة / Section:</strong> {section['section_number']}</div>
                <div><strong>المدرس / Instructor:</strong> {section['instructor']}</div>
                <div><strong>الفصل / Semester:</strong> {section['academic_year']} - {section['semester']}</div>
                <div><strong>النوع / Gender:</strong> {section['gender'].split(' / ')[0] if ' / ' in section['gender'] else section['gender']}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Progress metrics in columns
        metrics_col1, metrics_col2 = st.columns([1, 3])

        with metrics_col1:
            # Students progress
            students_icon = "✅" if section['students_count'] > 0 else "❌"
            st.markdown(f"""
            <div style="background: {'#e8f5e9' if section['students_count'] > 0 else '#ffebee'};
                        padding: 15px; border-radius: 8px; margin-bottom: 10px;">
                <h4 style="margin: 0;">{students_icon} الطلاب / Students</h4>
                <p style="font-size: 24px; font-weight: bold; margin: 5px 0;">
                    {section['students_count']} <small style="font-size: 14px;">({section['regular_count']} Regular)</small>
                </p>
            </div>
            """, unsafe_allow_html=True)

            # Overall progress
            overall_color = get_progress_color(section['overall_percentage'])
            overall_icon = get_progress_icon(section['overall_percentage'])
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, {overall_color}22 0%, {overall_color}44 100%);
                        padding: 15px; border-radius: 8px; border-left: 4px solid {overall_color};">
                <h4 style="margin: 0;">{overall_icon} الإنجاز الكلي / Overall</h4>
                <p style="font-size: 28px; font-weight: bold; margin: 5px 0; color: {overall_color};">
                    {section['overall_percentage']:.0f}%
                </p>
            </div>
            """, unsafe_allow_html=True)

        with metrics_col2:
            # Activities progress
            if activities:
                st.markdown("**تقدم الأنشطة / Activities Progress:**")

                # Create columns for activities (max 4 per row)
                activities_per_row = 4
                for i in range(0, len(activities), activities_per_row):
                    activity_cols = st.columns(min(activities_per_row, len(activities) - i))

                    for j, col in enumerate(activity_cols):
                        if i + j < len(activities):
                            activity = activities[i + j]
                            progress = section['activities_progress'].get(activity, {})
                            pct = progress.get('percentage', 0)
                            entered = progress.get('entered', 0)
                            total = progress.get('total', 0)
                            color = get_progress_color(pct)
                            icon = get_progress_icon(pct)

                            with col:
                                # Truncate activity name if too long
                                display_name = activity[:15] + "..." if len(activity) > 15 else activity
                                required_clos_count = progress.get('required_clos', 0)
                                st.markdown(f"""
                                <div style="background: #f8f9fa; padding: 10px; border-radius: 8px;
                                            border-left: 4px solid {color}; margin-bottom: 5px;">
                                    <p style="margin: 0; font-size: 12px; color: #666;" title="{activity}">{display_name}</p>
                                    <p style="margin: 0; font-size: 10px; color: #999;">({required_clos_count} CLOs)</p>
                                    <p style="margin: 2px 0; font-weight: bold; color: {color};">
                                        {icon} {pct:.0f}%
                                    </p>
                                    <p style="margin: 0; font-size: 11px; color: #888;">{entered}/{total}</p>
                                </div>
                                """, unsafe_allow_html=True)
            else:
                st.info("لا توجد أنشطة تقييم محددة / No assessment activities defined")

    st.markdown("---")

    # ===============================
    # Section 4: Summary Table & Export
    # ===============================
    st.markdown("### 4. Summary Table / جدول الملخص")

    # Create summary dataframe
    summary_data = []
    for section in sections_report_data:
        row = {
            'الشعبة / Section': section['section_number'],
            'المدرس / Instructor': section['instructor'],
            'الطلاب / Students': section['students_count'],
            'المنتظمين / Regular': section['regular_count'],
        }

        # Add activity columns
        for activity in activities:
            progress = section['activities_progress'].get(activity, {})
            pct = progress.get('percentage', 0)
            row[activity] = f"{pct:.0f}%"

        row['الإنجاز / Overall'] = f"{section['overall_percentage']:.0f}%"
        summary_data.append(row)

    if summary_data:
        df = pd.DataFrame(summary_data)
        st.dataframe(df, use_container_width=True, hide_index=True)

        # Export button
        st.markdown("<br>", unsafe_allow_html=True)
        col1, col2, col3 = st.columns([1, 1, 2])

        with col1:
            if EXCEL_AVAILABLE:
                excel_data = export_to_excel(sections_report_data, course_data.get('course_code', ''), activities_with_clos)
                st.download_button(
                    label="📥 Export to Excel",
                    data=excel_data,
                    file_name=f"section_progress_{course_data.get('course_code', '')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            else:
                st.warning("Excel export requires openpyxl package")

        with col2:
            # CSV export as fallback
            csv_data = df.to_csv(index=False).encode('utf-8-sig')
            st.download_button(
                label="📥 Export to CSV",
                data=csv_data,
                file_name=f"section_progress_{course_data.get('course_code', '')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv",
                use_container_width=True
            )

    # ===============================
    # Section 5: Legend
    # ===============================
    st.markdown("---")
    st.markdown("### Legend / دليل الألوان")

    legend_cols = st.columns(5)
    with legend_cols[0]:
        st.markdown("✅ **100%** - مكتمل / Complete")
    with legend_cols[1]:
        st.markdown("🔵 **75-99%** - جيد جداً / Very Good")
    with legend_cols[2]:
        st.markdown("🟡 **50-74%** - متوسط / Average")
    with legend_cols[3]:
        st.markdown("🟠 **1-49%** - منخفض / Low")
    with legend_cols[4]:
        st.markdown("🔴 **0%** - لم يبدأ / Not Started")
