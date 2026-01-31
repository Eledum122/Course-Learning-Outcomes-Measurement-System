"""
إدخال درجات طلاب الشعبة
Section Student Grades Entry
"""

import streamlit as st
import pandas as pd
import json
from pathlib import Path
from datetime import datetime
from io import BytesIO
from models.database import Database, UserRole
from utils.permissions import get_permissions_helper


def load_sections_data():
    """Load sections data"""
    sections_file = Path(__file__).parent.parent / 'data' / 'sections.json'
    if sections_file.exists():
        try:
            with open(sections_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {"sections": []}
    return {"sections": []}


def load_courses_data():
    """Load courses data"""
    courses_file = Path(__file__).parent.parent / 'data' / 'courses.json'
    if courses_file.exists():
        try:
            with open(courses_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {"courses": []}
    return {"courses": []}


def load_programs_data():
    """Load programs data"""
    programs_file = Path(__file__).parent.parent / 'data' / 'programs.json'
    if programs_file.exists():
        try:
            with open(programs_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {"programs": []}
    return {"programs": []}


def load_clos_data():
    """Load CLOs data"""
    clos_file = Path(__file__).parent.parent / 'data' / 'clos.json'
    if clos_file.exists():
        try:
            with open(clos_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {"clos": []}
    return {"clos": []}


def load_section_students_data():
    """Load section students data"""
    students_file = Path(__file__).parent.parent / 'data' / 'section_students.json'
    if students_file.exists():
        try:
            with open(students_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {"section_students": []}
    return {"section_students": []}


def load_student_grades_data():
    """Load student grades data"""
    grades_file = Path(__file__).parent.parent / 'data' / 'student_grades.json'
    if grades_file.exists():
        try:
            with open(grades_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {"student_grades": []}
    return {"student_grades": []}


def save_student_grades_data(data):
    """Save student grades data"""
    grades_file = Path(__file__).parent.parent / 'data' / 'student_grades.json'
    try:
        data['last_updated'] = datetime.now().isoformat()
        with open(grades_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except:
        return False


def get_section_students(section_id):
    """Get students for a specific section"""
    data = load_section_students_data()
    students = [s for s in data.get('section_students', []) if s.get('section_id') == section_id]
    return sorted(students, key=lambda x: x.get('seq', 0))


def get_course_clos(course_id):
    """Get CLOs for a course"""
    clos_data = load_clos_data()
    clos = [clo for clo in clos_data.get('clos', []) if clo.get('course_id') == course_id]
    return sorted(clos, key=lambda x: x.get('clo_code', ''))


def get_course_activities(course_id):
    """Get assessment activities for a course"""
    courses_data = load_courses_data()
    for course in courses_data.get('courses', []):
        if course.get('course_id') == course_id:
            return course.get('assessment_activities', [])
    return []


def get_course_data(course_id):
    """Get course data"""
    courses_data = load_courses_data()
    for course in courses_data.get('courses', []):
        if course.get('course_id') == course_id:
            return course
    return {}


def get_clos_activities_distribution(course_id):
    """Get CLOs distribution across activities"""
    courses_data = load_courses_data()
    for course in courses_data.get('courses', []):
        if course.get('course_id') == course_id:
            return course.get('clos_activities_distribution', {})
    return {}


def get_section_grades(section_id):
    """Get grades for a specific section"""
    data = load_student_grades_data()
    grades = {}
    for grade in data.get('student_grades', []):
        if grade.get('section_id') == section_id:
            student_id = grade.get('student_record_id')
            if student_id not in grades:
                grades[student_id] = {}
            key = f"{grade.get('activity_name')}_{grade.get('clo_code')}"
            grades[student_id][key] = grade.get('mark', 0)
    return grades


def save_section_grades(section_id, grades_dict):
    """Save grades for a section"""
    data = load_student_grades_data()

    # Remove existing grades for this section
    data['student_grades'] = [
        g for g in data['student_grades']
        if g.get('section_id') != section_id
    ]

    # Add new grades
    for student_id, student_grades in grades_dict.items():
        for key, mark in student_grades.items():
            parts = key.rsplit('_', 1)
            if len(parts) == 2:
                activity_name, clo_code = parts
                grade_record = {
                    "grade_id": f"grade_{section_id}_{student_id}_{activity_name}_{clo_code}",
                    "section_id": section_id,
                    "student_record_id": student_id,
                    "activity_name": activity_name,
                    "clo_code": clo_code,
                    "mark": mark,
                    "updated_at": datetime.now().isoformat()
                }
                data['student_grades'].append(grade_record)

    return save_student_grades_data(data)


def build_grades_structure(course_id):
    """
    Build the grades structure based on specifications table
    Returns: {activity_name: {clo_code: max_mark}}
    """
    clos_dist = get_clos_activities_distribution(course_id)
    structure = {}

    for activity_name, clo_marks in clos_dist.items():
        structure[activity_name] = {}
        for clo_code, mark in clo_marks.items():
            if mark > 0:
                structure[activity_name][clo_code] = mark

    return structure


def generate_grades_template(section_students, grades_structure, activities):
    """Generate Excel template for grades entry"""
    # Build columns
    columns = ['Seq', 'Student No', 'Student Name', 'Status']

    # Add activity/CLO columns
    activity_clo_columns = []
    for activity in activities:
        act_name = activity.get('assessment_task', '')
        if act_name in grades_structure:
            for clo_code, max_mark in grades_structure[act_name].items():
                # Include activity name in column to handle duplicate CLO/max combinations
                col_name = f"{act_name}\n{clo_code} ({max_mark:.0f})"
                activity_clo_columns.append({
                    'name': col_name,
                    'activity': act_name,
                    'clo': clo_code,
                    'max': max_mark
                })
                columns.append(col_name)

    columns.append('Total')

    # Build data
    data = []
    for student in section_students:
        row = {
            'Seq': student.get('seq'),
            'Student No': student.get('student_no'),
            'Student Name': student.get('student_name'),
            'Status': student.get('status')
        }
        for col in activity_clo_columns:
            row[col['name']] = ''
        row['Total'] = ''
        data.append(row)

    df = pd.DataFrame(data, columns=columns)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Grades')

        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Grades']

        # Style headers
        from openpyxl.styles import PatternFill, Font, Alignment

        # Activity header colors
        activity_colors = {
            0: 'FFF8DC',  # Cornsilk
            1: 'E6E6FA',  # Lavender
            2: 'E0FFFF',  # Light Cyan
            3: 'FAFAD2',  # Light Goldenrod
        }

        # Apply styles
        col_idx = 5  # Start after Status
        current_activity = None
        activity_index = 0

        for col_info in activity_clo_columns:
            if col_info['activity'] != current_activity:
                current_activity = col_info['activity']
                activity_index = (activity_index + 1) % len(activity_colors)

            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = PatternFill(start_color=activity_colors[activity_index],
                                   end_color=activity_colors[activity_index],
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            col_idx += 1

    output.seek(0)
    return output, activity_clo_columns


def show_student_grades(db: Database, user, lang: str):
    """Display Student Grades Entry page"""

    # Initialize grade version counter (used to force widget refresh)
    if 'grade_version' not in st.session_state:
        st.session_state.grade_version = 0

    # Check for reload flag - increment version to force new widget keys
    if st.session_state.get('force_reload_grades'):
        st.session_state.grade_version += 1
        st.session_state.force_reload_grades = False

    # تحديد صلاحية التعديل - منسق البرنامج ومنسق المقرر للاطلاع فقط
    can_edit = user.role not in [UserRole.PROGRAM_COORDINATOR, UserRole.COURSE_COORDINATOR]

    st.title("Student Grades / درجات الطلاب")
    if can_edit:
        st.caption("Enter and manage student grades for each section")
    else:
        st.caption("View student grades (Read-only) / عرض درجات الطلاب (للاطلاع فقط)")

    # Initialize permissions helper
    perm = get_permissions_helper(db, user)

    # Load sections and filter by permissions
    sections_data = load_sections_data()
    all_sections = sections_data.get('sections', [])
    sections = perm.filter_sections(all_sections)

    if not sections:
        st.warning("No sections assigned to you. Please contact your administrator.")
        return

    # Show messages
    if st.session_state.get('grades_message'):
        msg_type, msg_text = st.session_state.grades_message
        if msg_type == 'success':
            st.success(msg_text)
        elif msg_type == 'error':
            st.error(msg_text)
        elif msg_type == 'warning':
            st.warning(msg_text)
        st.session_state.grades_message = None

    # ===============================
    # Section 1: Select Section (with filters)
    # ===============================
    st.markdown("### 1. Select Section / اختر الشعبة")

    # Load additional data for filters and apply permissions
    programs_data = load_programs_data()
    all_programs = programs_data.get('programs', [])
    programs = perm.filter_programs(all_programs)

    courses_data = load_courses_data()
    all_courses = courses_data.get('courses', [])
    courses = perm.filter_courses(all_courses)

    # Row 1: Program and Course
    col1, col2 = st.columns(2)

    with col1:
        # Program filter (filtered by permissions)
        program_options = ["-- Select Program / اختر البرنامج --"]
        program_map = {}
        for p in programs:
            if p.get('is_active', True):
                display = f"{p.get('program_code', '')} - {p.get('program_name_en', '')}"
                program_options.append(display)
                program_map[display] = p.get('program_id')

        selected_program_display = st.selectbox(
            "Program / البرنامج *",
            program_options,
            key="grades_program_filter"
        )

        selected_program_id = None
        if selected_program_display != "-- Select Program / اختر البرنامج --":
            selected_program_id = program_map.get(selected_program_display)

    with col2:
        # Course filter (depends on program)
        if selected_program_id:
            program_courses = [c for c in courses if c.get('program_id') == selected_program_id]
            course_options = ["-- Select Course / اختر المقرر --"]
            course_map = {}
            for c in program_courses:
                display = f"{c.get('course_code', '')} - {c.get('course_title_en', '')}"
                course_options.append(display)
                course_map[display] = c.get('course_id')

            selected_course_display = st.selectbox(
                "Course / المقرر *",
                course_options,
                key="grades_course_filter"
            )

            selected_course_id = None
            if selected_course_display != "-- Select Course / اختر المقرر --":
                selected_course_id = course_map.get(selected_course_display)
        else:
            st.selectbox(
                "Course / المقرر *",
                ["-- Select Program First / اختر البرنامج أولاً --"],
                disabled=True,
                key="grades_course_filter_disabled"
            )
            selected_course_id = None

    # Row 2: Academic Year and Semester
    col3, col4 = st.columns(2)

    with col3:
        # Academic Year filter
        if selected_course_id:
            # Get available years for this course
            course_sections = [s for s in sections if s.get('course_id') == selected_course_id]
            available_years = sorted(list(set(s.get('academic_year', '') for s in course_sections if s.get('academic_year'))), reverse=True)

            year_options = ["-- Select Year / اختر السنة --"] + available_years
            selected_year = st.selectbox(
                "Academic Year / السنة الدراسية *",
                year_options,
                key="grades_year_filter"
            )
            if selected_year == "-- Select Year / اختر السنة --":
                selected_year = None
        else:
            st.selectbox(
                "Academic Year / السنة الدراسية *",
                ["-- Select Course First / اختر المقرر أولاً --"],
                disabled=True,
                key="grades_year_filter_disabled"
            )
            selected_year = None

    with col4:
        # Semester filter
        if selected_year and selected_course_id:
            # Get available semesters for this course and year
            year_sections = [s for s in sections if s.get('course_id') == selected_course_id and s.get('academic_year') == selected_year]
            available_semesters = list(set(s.get('semester', '') for s in year_sections if s.get('semester')))

            semester_display_options = ["-- Select Semester / اختر الفصل --"] + available_semesters
            selected_semester = st.selectbox(
                "Semester / الفصل الدراسي *",
                semester_display_options,
                key="grades_semester_filter"
            )
            if selected_semester == "-- Select Semester / اختر الفصل --":
                selected_semester = None
        else:
            st.selectbox(
                "Semester / الفصل الدراسي *",
                ["-- Select Year First / اختر السنة أولاً --"],
                disabled=True,
                key="grades_semester_filter_disabled"
            )
            selected_semester = None

    # Row 3: Section selection
    st.markdown("---")

    if selected_course_id and selected_year and selected_semester:
        # Filter sections based on all criteria
        filtered_sections = [
            s for s in sections
            if s.get('course_id') == selected_course_id
            and s.get('academic_year') == selected_year
            and s.get('semester') == selected_semester
        ]

        if filtered_sections:
            section_options = ["-- Select Section / اختر الشعبة --"]
            section_map = {}

            for s in filtered_sections:
                section_num = s.get('section_number', '')
                gender = s.get('gender', '').split(' / ')[0] if ' / ' in s.get('gender', '') else s.get('gender', '')
                beneficiary = ""
                if s.get('beneficiary_type') == 'external':
                    beneficiary = f" - 🏫 {s.get('beneficiary_department_en', '')[:20]}"

                option_text = f"Section {section_num} - {gender}{beneficiary}"
                section_options.append(option_text)
                section_map[option_text] = s

            col5, col6 = st.columns(2)

            with col5:
                selected_section_display = st.selectbox(
                    "Section / الشعبة *",
                    section_options,
                    key="grades_section_select"
                )

                selected_section = None
                if selected_section_display != "-- Select Section / اختر الشعبة --":
                    selected_section = section_map.get(selected_section_display)

            with col6:
                if selected_section:
                    course_data_info = get_course_data(selected_section.get('course_id'))
                    total_marks = course_data_info.get('total_course_marks', 100)
                    beneficiary_info = ""
                    if selected_section.get('beneficiary_type') == 'external':
                        beneficiary_info = f"\n**Beneficiary:** 🏫 {selected_section.get('beneficiary_college_en', '')} / {selected_section.get('beneficiary_department_en', '')}"

                    st.success(f"""
                    **Course:** {selected_section.get('course_code')}
                    **Section:** {selected_section.get('section_number')}
                    **Semester:** {selected_section.get('academic_year')} - {selected_section.get('semester_code')}
                    **Total Marks:** {total_marks}{beneficiary_info}
                    """)
        else:
            st.warning("No sections found for the selected criteria. Please check your selection.")
            selected_section = None
    else:
        st.info("💡 Please select Program, Course, Academic Year, and Semester to filter sections")
        selected_section = None

    if not selected_section:
        return

    section_id = selected_section.get('section_id')
    course_id = selected_section.get('course_id')

    # Get data
    section_students = get_section_students(section_id)
    course_clos = get_course_clos(course_id)
    course_activities = get_course_activities(course_id)
    course_data = get_course_data(course_id)
    grades_structure = build_grades_structure(course_id)
    clos_dist = get_clos_activities_distribution(course_id)

    # Validate data
    if not section_students:
        st.warning("No students in this section. Please add students first.")
        return

    if not course_activities:
        st.warning("No assessment activities defined. Please complete Stage 2 first.")
        return

    if not grades_structure:
        st.warning("CLOs distribution not defined. Please complete Stage 2.4 first.")
        return

    # Get saved grades
    saved_grades = get_section_grades(section_id)

    # Calculate totals
    total_marks = float(course_data.get('total_course_marks', 100))
    activity_marks = {}
    for activity in course_activities:
        act_name = activity.get('assessment_task')
        percentage = float(activity.get('percentage', 0))
        activity_marks[act_name] = (percentage / 100) * total_marks

    st.markdown("---")

    # ===============================
    # Section 2: Summary Header
    # ===============================
    # Count students by status
    status_counts = {'Regular': 0, 'Dropped': 0, 'Prohibited': 0, 'Incomplete': 0}
    for s in section_students:
        status = s.get('status', 'Regular')
        status_counts[status] = status_counts.get(status, 0) + 1

    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #2e7d32 0%, #1b5e20 100%);
                padding: 15px; border-radius: 10px; margin-bottom: 15px;">
        <h4 style="color: white; margin: 0;">
            {selected_section.get('course_code')} - Project &nbsp;&nbsp;
            Section: {selected_section.get('section_number')} - {selected_section.get('semester_code')} - {selected_section.get('gender', '').split(' / ')[0]}
        </h4>
        <p style="color: #c8e6c9; margin: 5px 0 0 0;">
            Total: {len(section_students)} | Regular: {status_counts['Regular']} | Dropped: {status_counts['Dropped']} | Prohibited: {status_counts['Prohibited']} | Incomplete: {status_counts['Incomplete']}
        </p>
    </div>
    """, unsafe_allow_html=True)

    # ===============================
    # Section 3: Upload/Download (only for users with edit permission)
    # ===============================
    if can_edit:
        st.markdown("### 2. Upload / Download")

        # Activity selection for import
        activity_names = [act.get('assessment_task', '') for act in course_activities if act.get('assessment_task', '') in grades_structure]
        activity_options = ["All Activities / جميع الأنشطة"] + activity_names

        col_activity, col_dl, col_ul = st.columns([1.5, 1, 1])

        with col_activity:
            selected_activity = st.selectbox(
                "Select Activity / اختر النشاط",
                activity_options,
                key="import_activity_select",
                help="اختر النشاط لاستيراد درجاته فقط، أو جميع الأنشطة لاستيراد كل الدرجات"
            )

        with col_dl:
            template_file, activity_clo_cols = generate_grades_template(
                section_students, grades_structure, course_activities
            )
            st.download_button(
                label="Download Template / تحميل القالب",
                data=template_file,
                file_name=f"grades_{selected_section.get('course_code')}_{selected_section.get('section_number')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )

        with col_ul:
            uploaded_file = st.file_uploader(
                "Upload Excel / رفع الملف",
                type=['xlsx', 'xls'],
                key="grades_upload"
            )
    else:
        uploaded_file = None

    if can_edit and uploaded_file:
        try:
            df = pd.read_excel(uploaded_file)
            st.markdown("**Preview:**")
            st.dataframe(df.head(5), use_container_width=True)

            # Show which activity will be imported
            if selected_activity == "All Activities / جميع الأنشطة":
                st.info("سيتم استيراد درجات **جميع الأنشطة** من الملف")
            else:
                st.info(f"سيتم استيراد درجات نشاط **{selected_activity}** فقط (الأنشطة الأخرى لن تتأثر)")

            if st.button("Import Grades / استيراد الدرجات", type="primary"):
                # Process uploaded file
                grades_data = load_student_grades_data()

                # Map student numbers to IDs
                student_map = {s.get('student_no'): s.get('student_record_id') for s in section_students}

                # Determine which activity to import
                import_all = selected_activity == "All Activities / جميع الأنشطة"
                target_activity = None if import_all else selected_activity

                imported_count = 0
                skipped_count = 0
                warnings_list = []  # Track grades that exceed max

                for _, row in df.iterrows():
                    student_no = str(row.get('Student No', '')).strip()
                    student_name = str(row.get('Student Name', '')).strip()
                    if student_no in student_map:
                        student_id = student_map[student_no]

                        for col in df.columns:
                            # Check for CLO columns
                            # New format: "Activity Name\nCLO (max)" e.g., "Lab Exam\nS4 (10)"
                            # Old format: "CLO\n(max)" or "CLO (max)" e.g., "S4\n(10)" or "S4 (10)"
                            col_str = str(col)
                            if '(' in col_str and col_str not in ['Seq', 'Student No', 'Student Name', 'Status', 'Total']:
                                # Parse column header
                                col_activity = None
                                clo_code = None
                                col_max = None

                                if '\n' in col_str:
                                    parts = col_str.split('\n')
                                    if len(parts) == 2 and '(' in parts[1]:
                                        # New format: "Activity Name\nCLO (max)"
                                        col_activity = parts[0].strip()
                                        clo_part = parts[1].strip()
                                        clo_code = clo_part.split('(')[0].strip()
                                        try:
                                            col_max = float(clo_part.split('(')[1].replace(')', '').strip())
                                        except:
                                            col_max = None
                                    else:
                                        # Old format: "CLO\n(max)"
                                        clo_code = parts[0].strip()
                                        try:
                                            col_max = float(parts[1].replace('(', '').replace(')', '').strip())
                                        except:
                                            col_max = None
                                else:
                                    # Old format: "CLO (max)"
                                    clo_code = col_str.split('(')[0].strip()
                                    try:
                                        col_max = float(col_str.split('(')[1].replace(')', '').strip())
                                    except:
                                        col_max = None

                                if not clo_code:
                                    continue

                                mark_val = row.get(col, 0)
                                if pd.notna(mark_val) and mark_val != '':
                                    mark_float = float(mark_val)

                                    # Find activity for this CLO
                                    matched_activity = None
                                    matched_max = None

                                    # If activity name is in column header, use it directly
                                    if col_activity and col_activity in grades_structure:
                                        if clo_code in grades_structure[col_activity]:
                                            matched_activity = col_activity
                                            matched_max = grades_structure[col_activity][clo_code]

                                    # Fall back to matching by CLO code and max mark
                                    if matched_activity is None:
                                        for act_name, clo_marks in grades_structure.items():
                                            if clo_code in clo_marks:
                                                act_max = clo_marks[clo_code]
                                                # Match by max mark from column header
                                                if col_max is not None and abs(act_max - col_max) < 0.01:
                                                    matched_activity = act_name
                                                    matched_max = act_max
                                                    break

                                    # If no exact match found, use first activity with this CLO
                                    if matched_activity is None:
                                        for act_name, clo_marks in grades_structure.items():
                                            if clo_code in clo_marks and clo_marks[clo_code] > 0:
                                                matched_activity = act_name
                                                matched_max = clo_marks[clo_code]
                                                break

                                    if matched_activity:
                                        # Skip if not the target activity (when specific activity selected)
                                        if not import_all and matched_activity != target_activity:
                                            skipped_count += 1
                                            continue

                                        # Track if grade exceeds max (don't cap it)
                                        if mark_float > matched_max:
                                            warnings_list.append({
                                                'student_no': student_no,
                                                'student_name': student_name,
                                                'clo': clo_code,
                                                'activity': matched_activity,
                                                'entered': mark_float,
                                                'max': matched_max
                                            })
                                            # Don't cap - keep original value to show in red

                                        key = f"{matched_activity}_{clo_code}"
                                        if student_id not in saved_grades:
                                            saved_grades[student_id] = {}
                                        saved_grades[student_id][key] = mark_float
                                        imported_count += 1

                if save_section_grades(section_id, saved_grades):
                    # Set flag to force reload grades on next run
                    st.session_state.force_reload_grades = True

                    # Build success/warning message
                    activity_msg = f"نشاط {target_activity}" if target_activity else "جميع الأنشطة"

                    if warnings_list:
                        # Build warning message - grades are shown in red, not auto-capped
                        warning_msg = f"⚠️ تم استيراد {imported_count} درجة لـ **{activity_msg}**. يوجد {len(warnings_list)} درجة تتجاوز الحد الأقصى (ستظهر باللون الأحمر):\n\n"
                        warning_msg += "| Student | CLO | Entered | Max |\n|---|---|---|---|\n"
                        for w in warnings_list[:10]:  # Show first 10 warnings
                            warning_msg += f"| {w['student_no']} | {w['clo']} | {w['entered']:.1f} | {w['max']:.1f} |\n"
                        if len(warnings_list) > 10:
                            warning_msg += f"\n... و {len(warnings_list) - 10} تحذيرات أخرى"
                        warning_msg += "\n\n**يرجى تصحيح الدرجات قبل الحفظ**"
                        st.session_state.grades_message = ('warning', warning_msg)
                        st.session_state.grades_warnings = warnings_list
                    else:
                        success_msg = f"تم استيراد {imported_count} درجة لـ **{activity_msg}** بنجاح!"
                        if skipped_count > 0:
                            success_msg += f" (تم تجاهل {skipped_count} درجة من أنشطة أخرى)"
                        st.session_state.grades_message = ('success', success_msg)
                    st.rerun()

        except Exception as e:
            st.error(f"Error reading file: {str(e)}")

    st.markdown("---")

    # ===============================
    # Section 4: Grades Table
    # ===============================
    st.markdown("### 3. Grades Table / جدول الدرجات")

    # Build activity-CLO structure for display
    activity_clo_list = []
    for activity in course_activities:
        act_name = activity.get('assessment_task', '')
        act_mark = activity_marks.get(act_name, 0)
        act_clos = []
        if act_name in grades_structure:
            for clo_code, max_mark in sorted(grades_structure[act_name].items()):
                act_clos.append({'clo': clo_code, 'max': max_mark})
        if act_clos:
            activity_clo_list.append({
                'activity': act_name,
                'mark': act_mark,
                'clos': act_clos
            })

    # Calculate column widths
    total_clo_cols = sum(len(a['clos']) for a in activity_clo_list)
    base_cols = 4  # Seq, Student No, Name, Status
    total_col = 1

    # Display table header - Activity level
    st.markdown("#### Activities and CLOs")

    # First row: Activity names with their total marks
    act_header_cols = st.columns([0.3, 0.8, 1.5, 0.6] + [len(a['clos']) * 0.5 for a in activity_clo_list] + [0.5])

    with act_header_cols[0]:
        st.markdown("**Seq**")
    with act_header_cols[1]:
        st.markdown("**Student No**")
    with act_header_cols[2]:
        st.markdown("**Student Name**")
    with act_header_cols[3]:
        st.markdown("**Status**")

    activity_colors = ['#fff3e0', '#e8f5e9', '#e3f2fd', '#fce4ec']
    for i, act_info in enumerate(activity_clo_list):
        with act_header_cols[4 + i]:
            color = activity_colors[i % len(activity_colors)]
            st.markdown(f"""
            <div style="background:{color};padding:5px;border-radius:5px;text-align:center;">
                <b>{act_info['activity']}</b><br>
                <span style="color:#666;">({act_info['mark']:.0f})</span>
            </div>
            """, unsafe_allow_html=True)

    with act_header_cols[-1]:
        st.markdown(f"""
        <div style="background:#e8f5e9;padding:5px;border-radius:5px;text-align:center;">
            <b>Total</b><br>
            <span style="color:#666;">({total_marks:.0f})</span>
        </div>
        """, unsafe_allow_html=True)

    # Second row: CLO codes with max marks
    clo_cols_def = [0.3, 0.8, 1.5, 0.6]
    for act_info in activity_clo_list:
        for _ in act_info['clos']:
            clo_cols_def.append(0.5)
    clo_cols_def.append(0.5)

    clo_header_cols = st.columns(clo_cols_def)

    col_idx = 4
    for i, act_info in enumerate(activity_clo_list):
        color = activity_colors[i % len(activity_colors)]
        for clo_info in act_info['clos']:
            with clo_header_cols[col_idx]:
                st.markdown(f"""
                <div style="background:{color};padding:3px;border-radius:3px;text-align:center;font-size:12px;">
                    <b>{clo_info['clo']}</b><br>
                    ({clo_info['max']:.0f})
                </div>
                """, unsafe_allow_html=True)
            col_idx += 1

    st.markdown("---")

    # Student rows with input fields
    all_grades = {}

    for student in section_students:
        student_id = student.get('student_record_id')
        student_no = student.get('student_no')
        student_name = student.get('student_name')
        status = student.get('status', 'Regular')

        if student_id not in all_grades:
            all_grades[student_id] = {}

        row_cols = st.columns(clo_cols_def)

        with row_cols[0]:
            st.markdown(f"**{student.get('seq')}**")

        with row_cols[1]:
            st.markdown(f"<small>{student_no}</small>", unsafe_allow_html=True)

        with row_cols[2]:
            name_short = student_name[:20] + "..." if len(student_name) > 20 else student_name
            st.markdown(f"<small>{name_short}</small>", unsafe_allow_html=True)

        with row_cols[3]:
            status_colors = {
                'Regular': '#28a745',
                'Dropped': '#ffc107',
                'Incomplete': '#17a2b8',
                'Prohibited': '#dc3545'
            }
            st.markdown(f"<small style='color:{status_colors.get(status, '#666')};'>{status}</small>", unsafe_allow_html=True)

        # Input fields for each CLO
        col_idx = 4
        row_total = 0
        has_exceeded = False

        for i, act_info in enumerate(activity_clo_list):
            for clo_info in act_info['clos']:
                key = f"{act_info['activity']}_{clo_info['clo']}"
                saved_val = saved_grades.get(student_id, {}).get(key, 0)
                max_val = float(clo_info['max'])

                with row_cols[col_idx]:
                    if status == 'Regular':
                        # Use version in key to force refresh after import/delete
                        widget_key = f"grade_v{st.session_state.grade_version}_{student_id}_{key}"

                        # Check if value exceeds max
                        is_exceeded = float(saved_val) > max_val

                        # Show red background if exceeded
                        if is_exceeded:
                            st.markdown(f"""
                            <style>
                                div[data-testid="stNumberInput"]:has(input[aria-label="{widget_key}"]) input {{
                                    background-color: #ffebee !important;
                                    border-color: #f44336 !important;
                                    color: #c62828 !important;
                                }}
                            </style>
                            """, unsafe_allow_html=True)
                            has_exceeded = True

                        val = st.number_input(
                            widget_key,
                            min_value=0.0,
                            max_value=1000.0,  # High limit to allow exceeded values
                            value=float(saved_val),
                            step=0.5,
                            key=widget_key,
                            label_visibility='collapsed',
                            disabled=not can_edit  # Read-only for program coordinators
                        )
                        all_grades[student_id][key] = val
                        row_total += min(val, max_val)  # Only count up to max for total
                    else:
                        st.markdown("<div style='background:#eee;height:32px;border-radius:3px;'></div>", unsafe_allow_html=True)
                        all_grades[student_id][key] = 0

                col_idx += 1

        # Total column
        with row_cols[-1]:
            if status == 'Regular':
                st.markdown(f"""
                <div style="background:#e8f5e9;padding:5px;border-radius:3px;text-align:center;font-weight:bold;">
                    {row_total:.1f}
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown("<div style='background:#eee;height:32px;border-radius:3px;text-align:center;'>-</div>", unsafe_allow_html=True)

    st.markdown("---")

    # Check for exceeded grades before save section
    exceeded_grades = []
    for student in section_students:
        student_id = student.get('student_record_id')
        student_no = student.get('student_no')
        if student.get('status') == 'Regular':
            for act_info in activity_clo_list:
                for clo_info in act_info['clos']:
                    key = f"{act_info['activity']}_{clo_info['clo']}"
                    grade_val = all_grades.get(student_id, {}).get(key, 0)
                    max_val = float(clo_info['max'])
                    if grade_val > max_val:
                        exceeded_grades.append({
                            'student_no': student_no,
                            'clo': clo_info['clo'],
                            'activity': act_info['activity'],
                            'entered': grade_val,
                            'max': max_val
                        })

    # ===============================
    # Section 5: Save/Delete Buttons (only for users with edit permission)
    # ===============================
    if can_edit:
        col1, col2, col3, col4 = st.columns([2, 1, 1, 1])

        with col1:
            if st.button("Save Grades / حفظ الدرجات", type="primary", use_container_width=True):
                if exceeded_grades:
                    # Show warning about exceeded grades
                    warning_msg = f"⚠️ لا يمكن الحفظ! يوجد {len(exceeded_grades)} درجة تتجاوز الحد الأقصى:\n\n"
                    warning_msg += "| Student | CLO | Entered | Max |\n|---|---|---|---|\n"
                    for w in exceeded_grades[:10]:
                        warning_msg += f"| {w['student_no']} | {w['clo']} | {w['entered']:.1f} | {w['max']:.1f} |\n"
                    if len(exceeded_grades) > 10:
                        warning_msg += f"\n... و {len(exceeded_grades) - 10} أخرى"
                    warning_msg += "\n\n**يرجى تصحيح الدرجات الحمراء أولاً**"
                    st.session_state.grades_message = ('error', warning_msg)
                    st.rerun()
                else:
                    if save_section_grades(section_id, all_grades):
                        st.session_state.grades_message = ('success', "Grades saved successfully! / تم حفظ الدرجات بنجاح!")
                        st.rerun()
                    else:
                        st.session_state.grades_message = ('error', "Error saving grades!")
                        st.rerun()

        with col2:
            if st.button("Reset / إعادة", use_container_width=True):
                st.rerun()

        with col3:
            # Delete button with confirmation
            if st.session_state.get('confirm_delete_grades') == section_id:
                if st.button("✓ Confirm / تأكيد", type="primary", use_container_width=True):
                    # Delete all grades for this section
                    data = load_student_grades_data()
                    original_count = len(data.get('student_grades', []))
                    data['student_grades'] = [
                        g for g in data.get('student_grades', [])
                        if g.get('section_id') != section_id
                    ]
                    deleted_count = original_count - len(data['student_grades'])
                    if save_student_grades_data(data):
                        # Set flag to force reload grades on next run
                        st.session_state.force_reload_grades = True
                        st.session_state.grades_message = ('success', f"تم حذف {deleted_count} درجة / Deleted {deleted_count} grades")
                        st.session_state.confirm_delete_grades = None
                        st.rerun()
            else:
                if st.button("🗑️ Delete / حذف", use_container_width=True):
                    st.session_state.confirm_delete_grades = section_id
                    st.rerun()

        with col4:
            if st.session_state.get('confirm_delete_grades') == section_id:
                if st.button("✗ Cancel / إلغاء", use_container_width=True):
                    st.session_state.confirm_delete_grades = None
                    st.rerun()
    else:
        # Show read-only notice for program coordinators
        st.info("🔒 أنت في وضع الاطلاع فقط - لا يمكنك تعديل الدرجات / You are in read-only mode - you cannot modify grades")

    # ===============================
    # Section 6: Summary
    # ===============================
    st.markdown("---")
    st.markdown("### 4. Summary / ملخص")

    # Calculate statistics
    regular_students = [s for s in section_students if s.get('status') == 'Regular']
    if regular_students:
        totals = []
        for student in regular_students:
            student_id = student.get('student_record_id')
            student_total = sum(all_grades.get(student_id, {}).values())
            totals.append(student_total)

        if totals:
            avg_total = sum(totals) / len(totals)
            max_total = max(totals)
            min_total = min(totals)
            passing = sum(1 for t in totals if t >= (total_marks * 0.6))

            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Average / المتوسط", f"{avg_total:.1f}")
            with col2:
                st.metric("Max / الأعلى", f"{max_total:.1f}")
            with col3:
                st.metric("Min / الأدنى", f"{min_total:.1f}")
            with col4:
                st.metric("Passing / الناجحون", f"{passing}/{len(regular_students)}")
