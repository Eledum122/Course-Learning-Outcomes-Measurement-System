"""
مولد ملف Excel لدرجات الطلاب
Grades Excel Exporter
"""

import os
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from models import Course, CourseSection


class GradesExcelExporter:
    """مولد ملف Excel لدرجات الطلاب"""

    def __init__(self, course: Course, section: CourseSection, language: str = 'ar'):
        """
        تهيئة المولد

        Args:
            course: كائن المقرر
            section: كائن الشعبة
            language: اللغة (ar أو en)
        """
        self.course = course
        self.section = section
        self.language = language
        self.is_rtl = (language == 'ar')

    def generate_excel(self, output_path: str):
        """
        توليد ملف Excel

        Args:
            output_path: مسار حفظ الملف
        """
        # إنشاء ملف Excel جديد
        wb = Workbook()
        ws = wb.active
        ws.title = "Grades" if self.language == 'en' else "الدرجات"

        # تعيين اتجاه الصفحة
        ws.sheet_view.rightToLeft = self.is_rtl

        # إنشاء الرأس
        self._create_header(ws)

        # إنشاء جدول الدرجات
        self._create_grades_table(ws)

        # حفظ الملف
        wb.save(output_path)
        print(f"[OK] Excel file generated: {output_path}")

    def _create_header(self, ws):
        """إنشاء رأس الجدول"""
        # معلومات المقرر
        ws['A1'] = "Course:" if self.language == 'en' else "المقرر:"
        ws['B1'] = f"{self.course.info.course_code} - {self.course.info.course_title}"
        ws['A1'].font = Font(bold=True, size=12)
        ws['B1'].font = Font(size=12)

        # معلومات الشعبة
        ws['A2'] = "Section:" if self.language == 'en' else "الشعبة:"
        ws['B2'] = f"{self.section.section_number} - {self.section.academic_year} - {self.section.semester.value}"
        ws['A2'].font = Font(bold=True, size=12)
        ws['B2'].font = Font(size=12)

        # أستاذ الشعبة
        ws['A3'] = "Instructor:" if self.language == 'en' else "أستاذ الشعبة:"
        instructor = self.section.section_instructor if self.section.section_instructor else ("Not Specified" if self.language == 'en' else "غير محدد")
        ws['B3'] = instructor
        ws['A3'].font = Font(bold=True, size=12)
        ws['B3'].font = Font(size=12)

    def _create_grades_table(self, ws):
        """إنشاء جدول الدرجات"""
        # بداية الجدول (الصف 5)
        start_row = 5

        # إنشاء رأس الجدول
        headers = []
        col_widths = []

        # أعمدة أساسية
        if self.is_rtl:
            headers.extend(['#', 'الرقم الجامعي', 'اسم الطالب', 'الحالة'])
            col_widths.extend([5, 15, 30, 12])
        else:
            headers.extend(['Seq', 'Student No', 'Student Name', 'Status'])
            col_widths.extend([5, 15, 30, 12])

        # أعمدة الأنشطة
        activities = sorted(self.course.activities, key=lambda a: a.name)
        activity_headers = []

        for activity in activities:
            if self.is_rtl:
                # عرض اسم النشاط والدرجة
                activity_header = f"{activity.name}\n({activity.mark})"
            else:
                activity_header = f"{activity.name}\n({activity.mark})"

            activity_headers.append(activity_header)
            col_widths.append(12)

        headers.extend(activity_headers)

        # عمود المجموع
        if self.is_rtl:
            headers.append(f"المجموع\n({self.course.info.total_mark})")
        else:
            headers.append(f"Total\n({self.course.info.total_mark})")
        col_widths.append(12)

        # كتابة رأس الجدول
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # تعيين عرض الأعمدة
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # كتابة بيانات الطلاب
        students = sorted(self.section.students, key=lambda s: s.student_id)

        for idx, student in enumerate(students, 1):
            row = start_row + idx
            col = 1

            # رقم تسلسلي
            self._write_cell(ws, row, col, idx, center=True)
            col += 1

            # الرقم الجامعي
            self._write_cell(ws, row, col, student.student_id, center=True)
            col += 1

            # اسم الطالب
            self._write_cell(ws, row, col, student.name, center=False)
            col += 1

            # الحالة
            status_text = student.status.value if hasattr(student.status, 'value') else str(student.status)
            if status_text == "Prohibited":
                status_color = "FFCDD2"  # أحمر فاتح
            elif status_text == "Dropped":
                status_color = "FFE082"  # أصفر فاتح
            else:
                status_color = "FFFFFF"  # أبيض

            self._write_cell(ws, row, col, status_text, center=True, bg_color=status_color)
            col += 1

            # درجات الأنشطة
            for activity in activities:
                activity_data = student.activity_marks.get(activity.name, {})
                # إذا كانت البيانات قاموس (درجات CLOs)، احسب المجموع
                if isinstance(activity_data, dict):
                    mark = sum(activity_data.values())
                else:
                    mark = activity_data if activity_data else 0.0

                self._write_cell(ws, row, col, mark if mark > 0 else '', center=True, number_format='0.0')
                col += 1

            # المجموع
            total = student.total_mark
            self._write_cell(ws, row, col, total, center=True, bold=True, number_format='0.0')

        # تجميد الصفوف والأعمدة العلوية
        ws.freeze_panes = ws.cell(row=start_row + 1, column=4 if self.is_rtl else 4)

    def _write_cell(self, ws, row, col, value, center=True, bold=False, bg_color=None, number_format=None):
        """كتابة خلية مع التنسيق"""
        cell = ws.cell(row=row, column=col)
        cell.value = value

        # الخط
        cell.font = Font(bold=bold, size=10)

        # المحاذاة
        if center:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='right' if self.is_rtl else 'left', vertical='center')

        # لون الخلفية
        if bg_color:
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

        # تنسيق الأرقام
        if number_format:
            cell.number_format = number_format

        # الحدود
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )


def generate_grades_excel(course: Course, section: CourseSection, output_path: str, language: str = 'ar'):
    """
    دالة مساعدة لتوليد ملف Excel للدرجات

    Args:
        course: كائن المقرر
        section: كائن الشعبة
        output_path: مسار حفظ الملف
        language: اللغة (ar أو en)

    Returns:
        True إذا نجح التوليد، False إذا فشل
    """
    try:
        exporter = GradesExcelExporter(course, section, language)
        exporter.generate_excel(output_path)
        return True
    except Exception as e:
        print(f"Error generating Excel file: {e}")
        import traceback
        traceback.print_exc()
        return False
